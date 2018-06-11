#!/usr/bin/python
# -*- coding:utf-8 -*-
'''
Created on 2016年11月22日

@author: 58
'''
from mysql.connector import connection
from openpyxl.reader.excel import load_workbook
from myutil import MyUtil
from operateDB import OperateDB
import re
import time
from string import zfill
from pip._vendor.requests.cookies import RequestsCookieJar
import requests
from mechanize._mechanize import Browser
from urllib import urlencode

class MyClass():
    '''
    classdocs
'''
    def test(self):
        add_company = ('''insert into listed_company(companyCode,companyShortName,  companyName,  companyEnlishName ,  
        ipoAddress,  aSharesCode ,  aSharesShortName,  aSharesIPODate, aSharesTotalCapital,  aSharesOutstandingCaptial,  
        bSharesCode,  bSharesShortName,  bSharesIPODate,  bSharesTotalCapital,  bSharesOutstandingCaptial,  area,
        province,  city,  trade,  website) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)''')
        data_company = ('000011', '深物业A', '深圳市物业发展(集团)股份有限公司', '''SHENZHEN PROPERTIES & RESOURCES DEVELOPMENT 
        (GROUP)CO., LTD''', '广东省深圳市人民南路国贸大厦39、42层', '11', '深物业A', '1992/3/30', '528,373,849', '175,831,365',
        '200011', '深物业B', '1992/3/30', '67,605,243', '67,605,243', '华南', '广东', '深圳市', 'K 房地产', 'www.szwuye.com.cn')
        cnx = connection.MySQLConnection(user='root', password='mysql',
                              host='127.0.0.1',
                              database='test')
        cursor = cnx.cursor()
        cursor.execute(add_company, data_company)
        cnx.commit()
        cursor.close()
        cnx.close()
    
    def fib(self, max):
        a, b = 0, 1
        while b < max:
            yield b
            a, b = b, a + b

    def f(self, x):
        return x * x
    
    def readXlsx(self):
        l = []
        wb = load_workbook(u"d:\\A股上市公司列表.xlsx")
        sheet = wb.active
        sum = 0
        for row in sheet.iter_rows():
            if sum == 0:
                sum = 1
                continue
            for num in range(len(row)):
                v = row[num].value
                l.append('%s' % v)
                print v
#                 if num == 7 or num == 12:
#                     ipoDate = str(row[num].value)[18:-7].replace(', ', '-')
#                     print ipoDate
            sum += 1
            print sum 
            print tuple(l)
          
def firstCharUpper(strr):
    srcStr = strr.strip()
    destStr = ''
    for i in range(len(srcStr)):
        if i == 0:
           destStr = srcStr[i].upper()
        else:
            destStr += srcStr[i]
    return destStr
            
def getWeather():
    print time.time()
    print MyUtil.getDatas("http://d1.weather.com.cn/sk_2d/101010100.html?_=", "http://www.weather.com.cn/weather1d/101010100.shtml")

def test1():
    cookie_dict = {'yfx_c_g_u_id_10000042':'_ck17051015271217105752979733547',
                   'VISITED_MENU':'%5B%229055%22%5D',
                   'yfx_f_l_v_t_10000042':'f_t_1494401232690__r_t_1494401232690__v_t_1494403377235__r_c_0'
    , 'sseMenuSpecial':'8527',
    'td_cookie':'18446744070125319008'}
    cookies = requests.utils.cookiejar_from_dict(cookie_dict, cookiejar=None, overwrite=True)
    
    s = requests.Session()
    s.cookies = cookies
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.133 Safari/537.36'}

    s.get('http://www.sse.com.cn/assortment/stock/list/info/company/index.shtml?COMPANY_CODE=601288', headers=headers)
#     s.get('http://sia.sseinfo.com/noc.gif?WS=10000042&RD=common&SWS=&SWSID=&SWSPID=&JSVER=20170109&TDT=&UC=_ck17051015271217105752979733547&LUC=&VUC=_vk1494401232690&FS=&RF=&PS=www.sse.com.cn&PU=%2Fassortment%2Fstock%2Flist%2Finfo%2Fcompany%2Findex.shtml%3FCOMPANY_CODE%3D600066&PT=&PER=0&PC=&PI=&LM=1494401232000&LG=zh-CN&CL=24&CK=1&SS=1366*768&SCW=371&SCH=638&SSH=879&FT=1494401232690&LT=1494401232690&DL=0&FL=0&CKT=HttpCookie&JV=0&AL=0&SY=windows%20nt%206.1&BR=chrome&TZ=-8&AU=&UN=&UID=&URT=&UA=&US=&TID=&MT=&FMSRC=same&MSRC=&MSCH=&EDM=&RC=0&SHPIC=&MID=1494401232690200&TT=%E5%85%AC%E5%8F%B8%E6%A6%82%E5%86%B5%20%7C%20%E4%B8%8A%E6%B5%B7%E8%AF%81%E5%88%B8%E4%BA%A4%E6%98%93%E6%89%80&CHK=126&SHT=sse.com.cn&RDM=0.320455976428041')
    r = s.get('http://query.sse.com.cn/commonQuery.do?jsonCallBack=jsonpCallback63654&isPagination=false&sqlId=COMMON_SSE_ZQPZ_GP_GPLB_C&productid=601288&_=1494238148603', headers=headers)
    print r
   
def test2():
    br = Browser()  # Create a browser
    br.open('http://www.sse.com.cn/assortment/stock/list/info/company/index.shtml?COMPANY_CODE=601288') 
    print br.response().read()

if __name__ == '__main__':
    print u'\u9ec4\u5c71\u65c5\u6e38/HSTD'.encode('utf8')
    mc = MyClass()
    test1()
    
#     mc.readXlsx()
    l = ['adb', 'abc']
    print tuple(l)
    print len('ANHUI WANWEI UPDATED HIGH-TECH MATERIAL INDUSTRY COMPANY LIMITED')

    print l.sort()
    print l
    ll = list(MyUtil.indexs.split(','))
    print ll
    print map(lambda x:'get' + x, map(firstCharUpper, ll))
    
    downLoadSZSE = raw_input("download szse xls?Y|N")
#     print downLoadSZSE
    m = re.search('^[Y|y][E|e]{0,}[S|s]{0,}', downLoadSZSE)
    if m  :
        print m.group()
    else:
        print 'ddd'
    getWeather()
