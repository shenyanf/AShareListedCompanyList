#!/usr/bin/python
# -*- coding:utf-8 -*-
'''
Created on 2016年11月22日

@author: 58
'''
from mysql.connector import connection
import xlrd
import urllib
from openpyxl.reader.excel import load_workbook
import datetime
from myutil import MyUtil
from operateDB import OperateDB

class MyClass():
    '''
    classdocs
'''
    def test(self):
        add_company = ("insert into listed_company(companyCode,companyShortName,  companyName,  companyEnlishName ,  ipoAddress,  aSharesCode ,  aSharesShortName,  aSharesIPODate,\
  aSharesTotalCapital,  aSharesOutstandingCaptial,  bSharesCode,  bSharesShortName,  bSharesIPODate,  bSharesTotalCapital,  bSharesOutstandingCaptial,  area,\
  province,  city,  trade,  website) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)")
        data_company = ('000011', '深物业A', '深圳市物业发展(集团)股份有限公司', 'SHENZHEN PROPERTIES & RESOURCES DEVELOPMENT (GROUP)CO., LTD', '广东省深圳市人民南路国贸大厦39、42层', '11', '深物业A', '1992/3/30', '528,373,849', '175,831,365', '200011', '深物业B', '1992/3/30', '67,605,243', '67,605,243', '华南', '广东', '深圳市', 'K 房地产', 'www.szwuye.com.cn')
        cnx = connection.MySQLConnection(user='root', password='mysql',
                              host='127.0.0.1',
                              database='test')
        cursor = cnx.cursor()
        cursor.execute(add_company, data_company)
        cnx.commit()
        cursor.close()
        cnx.close()
    
    def fib(self, max):
        a, b = 0, 1
        while b < max:
            yield b
            a, b = b, a + b

    def f(self, x):
        return x * x
    
    def readXlsx(self):
        l = []
        wb = load_workbook(u"d:\\A股上市公司列表.xlsx")
        sheet = wb.active
        sum = 0
        for row in sheet.iter_rows():
            if sum == 0:
                sum = 1
                continue
            for num in range(len(row)):
                v = row[num].value
                l.append('%s' % v)
                print v
#                 if num == 7 or num == 12:
#                     ipoDate = str(row[num].value)[18:-7].replace(', ', '-')
#                     print ipoDate
            sum += 1
            print sum 
            print tuple(l)
            
if __name__ == '__main__':
    print u'\u9ec4\u5c71\u65c5\u6e38/HSTD'.encode('utf8')
    mc = MyClass()
#     mc.readXlsx()
    l = [1, 2, 3, 4, 5, 5, 6]
    print tuple(l)
    print len('ANHUI WANWEI UPDATED HIGH-TECH MATERIAL INDUSTRY COMPANY LIMITED')
    odb = OperateDB()
    stockInfos = odb.selectFromCompany(MyUtil.selectCompany.replace('%s', MyUtil.indexs))
    for info in stockInfos:
        print info
#         
        
#         for j in range(len(stockInfos[info])):
#             print stockInfos[info][j]
