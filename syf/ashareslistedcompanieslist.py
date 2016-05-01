# -*- coding: gbk -*- 
'''
Created on 2016年4月20日

@author: shenyf
'''
import xlrd
from openpyxl.workbook import Workbook
import urllib
from syf.achievesselistedcompanyinfo import AchieveSSEStockInfo
from openpyxl.compat import range
from time import sleep

class ASharesListedCompaniesList:
    '''A股上市公司列表
    目前从深圳交易所下载下来的《上市公司列表》格式就有问题，[无语...]，因此xls2xlsx 还有向深圳《上市公司列表》中
    追加上海交易所上市公司信息都是不可能的，只能先手动下载深圳《上市公司列表》，然后抓取上海上市公司信息，然后手动合并[无语中...]'''
    
    # file absolute path
    filePath = u"d:\\上证所A股上市公司列表.xlsx"
    # 上海交易所上市公司起始代码为600000
    startStockNumber = 600000
    # xlsx 标题
    __indexName = [u'公司代码', u'公司简称', u'公司全称', u'英文名称', u'注册地址', u'A股代码', u'A股简称', u'A股上市日期', u'A股总股本', u'A股流通股本',
                   u'B股代码', u'B股简称', u'B股上市日期', u'B股总股本', u'B股流通股本', u'地 区', u'省 份', u'城 市', u'所属行业', u'公司网址']
    
    def appendSSEStocks(self, lastStockNumber):
        wb = Workbook()  # load_workbook(filename=self.filePath)
        sheet = wb.active
        
        # 写标题
        for i in range(self.__indexName.__len__()):
            _ = sheet.cell(column=i + 1, row=1, value=self.__indexName[i])
            
        row = sheet.max_row + 1
        for i in range(self.startStockNumber, lastStockNumber + 1):
            a = AchieveSSEStockInfo(i)
            sleep(1)
            
#             print a.getStatus()
            if not a.getStatus():
                continue
            
            for j in range(a.__public__.__len__()):
                m = a.__public__[j]
                f = getattr(a, m)
                print m
                print f()
                _ = sheet.cell(column=j + 1, row=row, value="%s" % f())
            row = row + 1
            
            # 每获取一个上市公司完整信息就写入xlsx，避免占用过大内存
            wb.save(filename=self.filePath)
    
    def downloadSZSEASharesListedCompaniesList(self):
        '''下载深圳交易所上市公司信息.'''
        dls = r"http://www.szse.cn/szseWeb/ShowReport.szse?SHOWTYPE=EXCEL&CATALOGID=1110&tab1PAGENUM=1&ENCODE=1&TABKEY=tab1"
        urllib.urlretrieve(dls, self.filePath)
        
    def storeASharesListedCompanies2XLS(self, lastStockNumber):
#         self.downloadSZSEASharesStocksInfo()
#         self.xls2xlsx()
        self.appendSSEStocks(lastStockNumber=lastStockNumber)

    def xls2xlsx(self):
        # first open using xlrd
        book = xlrd.open_workbook(self.filePath)
        index = 0
        nrows, ncols = 0, 0
        while nrows * ncols == 0:
            sheet = book.sheet_by_index(index)
            nrows = sheet.nrows
            ncols = sheet.ncols
            index += 1
    
        # prepare a xlsx sheet
        book1 = Workbook()
        sheet1 = book1.active
    
        for row in xrange(1, nrows):
            for col in xrange(1, ncols):
                sheet1.cell(row=row, column=col).value = sheet.cell_value(row, col)
        
        book1.save(filename=u'd:\\A.xlsx')
        return book1
        
if __name__ == '__main__':
    s = ASharesListedCompaniesList()
    s.storeASharesListedCompanies2XLS(lastStockNumber=600001)
