# -*- coding: utf-8 -*- 
'''
Created on 2016年10月08日
@author: shenyanf
'''

import urllib
from achievestockinfo import AchieveSSEStockInfo
from openpyxl.reader.excel import load_workbook
from openpyxl.styles.borders import Border, Side
from operateDB import OperateDB
import datetime
from openpyxl.workbook.workbook import Workbook
import os
import code
from syf.myutil import MyUtil
import pandas as pd

class ASharesStocks:
    '''
    下载深交所上市公司信息，抓取上交所上市公司信息，入库后，导出到xlsx中
    '''
    
    fileName = u'A股上市公司列表'
    # file absolute path
    filePath = u"d:\\" + fileName + ".xlsx"
    downloadFilePath = filePath.replace('xlsx', 'xls')
    
    # xlsx 标题
    __indexName = [u'公司代码', u'公司简称', u'公司全称', u'英文名称', u'注册地址', u'A股代码', u'A股简称', u'A股上市日期', u'A股总股本',
                   u'A股流通股本', u'B股代码', u'B股简称', u'B股上市日期', u'B股总股本', u'B股流通股本', u'地 区', u'省 份', u'城 市',
                    u'所属行业', u'公司网址']
    
    def appendSSEStocks(self):
        '''
        向filePath指定表格中追加上交所上市公司信息
        '''
        wb = load_workbook(self.filePath)
        sheet = wb.active
#         print sheet.max_row
        # 设置单元格所有框线
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))   
#         # 写标题
#         for i in range(self.__indexName.__len__()):
#             _ = sheet.cell(column=i + 1, row=1, value=self.__indexName[i])
        # 当前表格的信息行数
        row = sheet.max_row
        for i in self.getSSEAShareListCompanyCode():
            a = AchieveSSEStockInfo(i)
             
            if not a.getStatus():
                continue
             
            for j in range(len(a.methodList)):
                m = a.methodList[j]
                f = getattr(a, m)
                _ = sheet.cell(column=j + 1, row=row, value="%s" % f())
                sheet.cell(column=j + 1, row=row).border = thin_border
            row = row + 1
            
            if int(i) % 100 == 0:
                wb.save(self.filePath)
        wb.save(self.filePath)
            
    def downloadSZSEAShares(self):
        '''
        下载深圳交易所上市公司信息.保存到 self.downloadFilePath中
        @return download success or not
        '''
        dls = 'http://www.szse.cn/szseWeb/ShowReport.szse?SHOWTYPE=EXCEL&CATALOGID=1110&tab1PAGENUM=1&ENCODE=1&' + \
        'TABKEY=tab1'
        try:
            if os.path.exists(self.downloadFilePath):
                os.remove(self.downloadFilePath)
            urllib.urlretrieve(dls, self.downloadFilePath)

            # 下载后直接零存为xlsx格式
            self.__htmlTable2xlsx()
        except Exception, e:
            print e
            return False
        else:
            return True
        
    def __htmlTable2xlsx(self):
        '''
        直接下载的深交所上市公司信息是html table格式的，使用pandas可以直接保存到xlsx中 
        '''
        with open(self.downloadFilePath, 'r') as f:
            dataFrames = pd.read_html(f.read())

        # 只有一个sheet
        df1 = dataFrames[0]
    
        writer = pd.ExcelWriter(self.filePath, engine='xlsxwriter')
        df1.to_excel(writer , encoding="utf-8")
        writer.save()
        
        f.close()
        writer.close()
        
        # 删除下载的不可用的xls文件
        if os.path.exists(self.downloadFilePath):
            os.remove(self.downloadFilePath)
        
    def __mergeURL(self, pageNum, firstPage=True):
        """
        @param pageNum: 第几页
        @param firstPage:  是否是第一页，为true时，url不需要拼接endPage
        @return:    合成的url
        """
        return 'http://query.sse.com.cn/security/stock/getStockListData2.do?&jsonCallBack=jsonpCallback53818&is' + \
        'Pagination=true&stockCode=&csrcCode=&areaName=&stockType=1&pageHelp.cacheSize=1&pageHelp.beginPage=' + \
        pageNum + '&pageHelp.pageSize=25' + (firstPage and '&pageHelp.endPage=' + pageNum + '1' or '') + \
        '&pageHelp.pageNo=' + pageNum + '&_='
    
    def getSSEAShareListCompanyCode(self):
        """
        获取上交所所有上市公司股票代码
        @return: list, all sse list company code, get from website
        """
        
        print 'get all sse list company code'
        
        l = []
        totalCompany = 0
        referer = r'http://www.sse.com.cn/assortment/stock/list/share/'
#         第一次获取表格首页信息及总页数等信息
        jsonCallBackPythonObjData = MyUtil.getDatas(self.__mergeURL('1'), 'pageHelp', referer)
        if not jsonCallBackPythonObjData:
            return None
        else:
            # 总页数
            pageCount = jsonCallBackPythonObjData['pageCount']
            # 每页多少条记录
            pageSize = jsonCallBackPythonObjData['pageSize']
            # 总共多少上市公司
            total = jsonCallBackPythonObjData['total']

        for i in range(1, pageCount + 1):
            # 获取第i页上市公司信息
            jsonCallBackPythonObjData = MyUtil.getDatas(self.__mergeURL(str(i), False), 'pageHelp', referer)
            if totalCompany == total:
                break;
            datas = jsonCallBackPythonObjData['data']
            for a in datas:
                companyCode = a['COMPANY_CODE']
                l.append(companyCode)
                totalCompany += 1
        
        if len(l) != total:
            print 'The total does not match the actual'

        print l
        return l
    
    def sseCompanyStore2DB(self):
        '''
        store sse list company which doesn't exist in database to database
        @return: if store success, True, other, False
        '''
        print 'achieve sse list company info, and store to database'
        odb = OperateDB()
        try:
            # list of company code, those listed but not store in database
            resList = list(set(self.getSSEAShareListCompanyCode()) ^ set(odb.selectSSEAllCompanyCode()))
            print "need fetch :%s" % resList
            for i in resList:
                stockinfo = AchieveSSEStockInfo(i)
                    
                if not stockinfo.getStatus():
                    continue
                l = "" 
                # change company info  to a string
                l = '|'.join(stockinfo.allCompanyInfo())
            
                odb.addCompany(MyUtil.addCompany, tuple(item for item in l.split('|')))
        except Exception, e:
            print e
            return False
        else:
            print 'store to db successful'
            return True
        
    def szseCompanyStore2DB(self):
        '''
        like @method sseCompanyStore2DB 
        store szse list company to database, before use this method you need call method downloadSZSEAShares and 
        manually change excel type to xlsx 
        @return: if store success, True, other, False
        '''
        odb = OperateDB()
        resList = list(set(odb.selectSZSEAllCompanyCode()))
#         print len(resList)
        wb = load_workbook(self.filePath)
        sheet = wb.active
        total = 0
        for row in sheet.iter_rows():
            l = []
            # 不是六位在前面填充0，例如：1格式化为000001
            code = str(row[0].value).zfill(6)
            
            try:
                # check first column value is a valid company code and not in database 
                if int(code) and code not in resList:
                    # append companyCode to l
                    l.append(code)
                    for num in range(1, len(row)):
                        v = row[num].value
                        # '%s' % v, change datatime.datatime(xxxx,xx,xx,xx,xx) and float and etc to a string
                        if num == 3 or num == 19:
                            l.append(MyUtil.delChinese('%s' % v))
                        else:
                            l.append('%s' % v)
                    
                    print 'add szse company %s' % row[0].value
                    odb.addCompany(MyUtil.addCompany, tuple(l))
                    total += 1
            except Exception, e:
                pass
            
        print 'add %d companies' % total
        return True
        
    def store2xlsx(self):
        '''
        select company info and store to xlsx
        '''
        odb = OperateDB()
        
        # current date
        datee = datetime.datetime.now().strftime("%Y-%m-%d")
        
        fileAbsPath = self.filePath.replace('.xlsx', datee + '.xlsx')
        
        wb = Workbook()
        sheet = wb.active
#         print sheet.max_row
        # 设置单元格所有框线
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(
style='thin'))   
        # 写标题
        for i in range(self.__indexName.__len__()):
            _ = sheet.cell(column=i + 1, row=1, value=self.__indexName[i])
        
        stockInfos = odb.selectFromCompany(MyUtil.selectCompany.replace('%s', MyUtil.indexs))
        
        keyList = stockInfos.keys()
        # 按照公司代码排序
        keyList.sort()
#         print keyList

        # 第一行为标题，从第二行开始为实际数据
        rowNum = 2
        for i in keyList:
            for j in range(len(stockInfos[i])):
                _ = sheet.cell(column=j + 1, row=rowNum, value=stockInfos[i][j])
                sheet.cell(column=j + 1, row=rowNum).border = thin_border
            rowNum += 1
            
            # 每100条保存一下
            if int(i) % 100 == 0:
                wb.save(fileAbsPath)
        # 最后保存一下
        wb.save(fileAbsPath)
        
if __name__ == '__main__':
    s = ASharesStocks()
    # 下载深圳交易所上市公司信息, 需要手动打开并另存为xlsx
    s.downloadSZSEAShares()
    print 'Download finish. file full path:%s' % s.filePath
    # 初始化辅助类
    MyUtil()
    # 获取上交所上市公司并入库及获取深交所上市公司并入库
    if  s.sseCompanyStore2DB() and  s.szseCompanyStore2DB():
        # 数据库中数据保存到xlsx中
        s.store2xlsx()
        print 'all A share listed company already store to database and export to %s' % s.filePath 
